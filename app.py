"""
Test Scenario & Coverage Generator - Web Application
A web interface for generating structured test scenarios from code changes.
"""

from flask import Flask, render_template, request, jsonify, Response, send_file
from flask_cors import CORS
import io
import json
import os
from dotenv import load_dotenv
import traceback
from datetime import datetime

import boto3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from src.git_analyzer import GitHubPRAnalyzer
from src.code_analyzer import CodeAnalyzer
from src.test_generator import TestScenarioGenerator
from src.excel_processor import ExcelProcessor, ExcelParseError
from src.excel_mapper import ExcelMapper
from src.jira_client import ZephyrScaleClient

# Load environment variables
import pathlib
env_path = pathlib.Path(__file__).parent / '.env'
load_dotenv(dotenv_path=env_path, override=True)

# Initialize Flask app
app = Flask(__name__)
CORS(app)

app.config['SECRET_KEY'] = os.getenv('FLASK_SECRET_KEY', 'dev-secret-key-change-in-production')
app.config['MAX_CONTENT_LENGTH'] = 32 * 1024 * 1024  # 32MB max (Excel uploads)

# Shared Bedrock client — initialised once, reused across requests
_bedrock_client = boto3.client(
    "bedrock-runtime",
    region_name=os.getenv("AWS_REGION", "us-east-1"),
    aws_access_key_id=os.getenv("AWS_ACCESS_KEY_ID"),
    aws_secret_access_key=os.getenv("AWS_SECRET_ACCESS_KEY"),
    aws_session_token=os.getenv("AWS_SESSION_TOKEN"),
)


@app.before_request
def require_basic_auth():
    if request.path == '/health':
        return None

    app_username = os.getenv('APP_USERNAME', '')
    app_password = os.getenv('APP_PASSWORD', '')

    if not app_username or not app_password:
        return None

    auth = request.authorization
    if not auth or auth.username != app_username or auth.password != app_password:
        return Response(
            'Authentication required.',
            401,
            {'WWW-Authenticate': 'Basic realm="Test Scenario Generator"'}
        )


@app.route('/')
def index():
    return render_template('index.html')


# ─────────────────────────────────────────────
# Analysis endpoints
# ─────────────────────────────────────────────

@app.route('/api/analyze-pr', methods=['POST'])
def analyze_pr():
    """
    Analyse a GitHub pull request and generate structured test cases.

    JSON payload: { "pr_url": "...", "generate_code": false }
    """
    try:
        data = request.get_json()
        if not data or 'pr_url' not in data:
            return jsonify({'success': False, 'error': 'Missing PR URL'}), 400

        pr_url = data['pr_url'].strip()
        generate_code = data.get('generate_code', False)

        if 'github.com' not in pr_url:
            return jsonify({'success': False, 'error': 'Invalid GitHub PR URL. Must contain github.com'}), 400

        parts = pr_url.split('/')
        if len(parts) < 7:
            return jsonify({'success': False, 'error': 'Invalid GitHub PR URL format'}), 400

        owner = parts[-4]
        repo = parts[-3]
        try:
            pr_number = int(parts[-1])
        except ValueError:
            return jsonify({'success': False, 'error': 'Invalid PR number'}), 400

        # Step 1: Fetch from GitHub
        gh_analyzer = GitHubPRAnalyzer(os.getenv('GITHUB_TOKEN'))
        try:
            diff_text = gh_analyzer.get_pr_diff(owner, repo, pr_number)
            pr_info = gh_analyzer.get_pr_info(owner, repo, pr_number)
        except Exception as e:
            error_msg = str(e)
            if '401' in error_msg:
                return jsonify({'success': False, 'error': 'GitHub authentication failed. Add GITHUB_TOKEN to .env'}), 401
            elif '404' in error_msg:
                return jsonify({'success': False, 'error': 'PR not found. Check the URL and repository access'}), 404
            return jsonify({'success': False, 'error': f'GitHub API error: {error_msg}'}), 500

        # Step 2: Analyse code changes (local)
        code_analyzer = CodeAnalyzer()
        parsed_diff = code_analyzer.parse_diff(diff_text)
        change_types = code_analyzer.identify_change_types(parsed_diff)
        diff_summary = code_analyzer.generate_summary(parsed_diff)

        # Step 3: Single Bedrock call → structured test cases
        test_generator = TestScenarioGenerator()
        try:
            structured_test_cases = test_generator.generate_structured_test_cases(
                diff_summary=diff_summary,
                parsed_diff=parsed_diff,
                change_types=change_types,
                pr_context=pr_info,
            )
        except Exception as e:
            error_msg = str(e)
            if 'throttlingexception' in error_msg.lower() or 'toomanyrequests' in error_msg.lower():
                return jsonify({'success': False, 'error': 'AWS Bedrock throttled. Please retry in a moment.'}), 429
            elif 'accessdeniedexception' in error_msg.lower() or 'is not authorized' in error_msg.lower():
                return jsonify({'success': False, 'error': 'AWS credentials invalid or lack Bedrock access.'}), 401
            return jsonify({'success': False, 'error': f'AI generation error: {error_msg}'}), 500

        # Step 4: Optionally generate test code
        test_code = None
        if generate_code:
            try:
                test_code = test_generator.generate_automated_test_code(
                    structured_test_cases=structured_test_cases,
                    language="python",
                    framework="pytest",
                )
            except Exception as e:
                test_code = f"Error generating test code: {str(e)}"

        # Step 5: Build file analyses
        file_analyses = [
            {
                'file_path': f['file_path'],
                'additions': len(f['additions']),
                'deletions': len(f['deletions']),
                'has_additions': len(f['additions']) > 0,
                'has_deletions': len(f['deletions']) > 0,
            }
            for f in parsed_diff[:10]
        ]

        return jsonify({
            'success': True,
            'data': {
                'pr_info': {
                    'title': pr_info['title'],
                    'author': pr_info['author'],
                    'base_branch': pr_info['base_branch'],
                    'head_branch': pr_info['head_branch'],
                    'state': pr_info['state'],
                },
                'summary': {
                    'total_files': len(parsed_diff),
                    'total_additions': sum(len(f['additions']) for f in parsed_diff),
                    'total_deletions': sum(len(f['deletions']) for f in parsed_diff),
                },
                'file_analyses': file_analyses,
                'change_types': change_types,
                'structured_test_cases': structured_test_cases,
                'test_code': test_code,
                'generated_at': datetime.now().isoformat(),
            }
        })

    except Exception as e:
        app.logger.error("analyze_pr error: %s", traceback.format_exc())
        return jsonify({'success': False, 'error': f'Unexpected error: {str(e)}'}), 500


@app.route('/api/analyze-diff', methods=['POST'])
def analyze_diff():
    """
    Analyse a manually pasted git diff and generate structured test cases.

    JSON payload: { "diff_text": "...", "generate_code": false }
    """
    try:
        data = request.get_json()
        if not data or 'diff_text' not in data:
            return jsonify({'success': False, 'error': 'Missing diff text'}), 400

        diff_text = data['diff_text'].strip()
        generate_code = data.get('generate_code', False)

        if not diff_text:
            return jsonify({'success': False, 'error': 'Diff text is empty'}), 400

        # Analyse code changes (local)
        code_analyzer = CodeAnalyzer()
        parsed_diff = code_analyzer.parse_diff(diff_text)
        change_types = code_analyzer.identify_change_types(parsed_diff)
        diff_summary = code_analyzer.generate_summary(parsed_diff)

        # Single Bedrock call → structured test cases
        test_generator = TestScenarioGenerator()
        try:
            structured_test_cases = test_generator.generate_structured_test_cases(
                diff_summary=diff_summary,
                parsed_diff=parsed_diff,
                change_types=change_types,
                pr_context=None,
            )
        except Exception as e:
            error_msg = str(e)
            if 'throttlingexception' in error_msg.lower() or 'toomanyrequests' in error_msg.lower():
                return jsonify({'success': False, 'error': 'AWS Bedrock throttled. Please retry in a moment.'}), 429
            elif 'accessdeniedexception' in error_msg.lower() or 'is not authorized' in error_msg.lower():
                return jsonify({'success': False, 'error': 'AWS credentials invalid or lack Bedrock access.'}), 401
            return jsonify({'success': False, 'error': f'AI generation error: {error_msg}'}), 500

        # Optionally generate test code
        test_code = None
        if generate_code:
            try:
                test_code = test_generator.generate_automated_test_code(
                    structured_test_cases=structured_test_cases,
                    language="python",
                    framework="pytest",
                )
            except Exception as e:
                test_code = f"Error generating test code: {str(e)}"

        file_analyses = [
            {
                'file_path': f['file_path'],
                'additions': len(f['additions']),
                'deletions': len(f['deletions']),
            }
            for f in parsed_diff[:10]
        ]

        return jsonify({
            'success': True,
            'data': {
                'summary': {
                    'total_files': len(parsed_diff),
                    'total_additions': sum(len(f['additions']) for f in parsed_diff),
                    'total_deletions': sum(len(f['deletions']) for f in parsed_diff),
                },
                'file_analyses': file_analyses,
                'change_types': change_types,
                'structured_test_cases': structured_test_cases,
                'test_code': test_code,
                'generated_at': datetime.now().isoformat(),
            }
        })

    except Exception as e:
        app.logger.error("analyze_diff error: %s", traceback.format_exc())
        return jsonify({'success': False, 'error': f'Error: {str(e)}'}), 500


# ─────────────────────────────────────────────
# Excel download endpoints
# ─────────────────────────────────────────────

@app.route('/api/download-test-cases-excel', methods=['POST'])
def download_test_cases_excel():
    """
    Convert structured test cases JSON to a formatted Excel workbook.

    JSON payload: { "test_cases": [...] }
    Returns: .xlsx file download
    """
    try:
        data = request.get_json()
        test_cases = data.get('test_cases', []) if data else []

        if not test_cases:
            return jsonify({'success': False, 'error': 'No test cases provided'}), 400

        wb = Workbook()
        ws = wb.active
        ws.title = 'Test Cases'

        header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        alt_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')

        headers = ['ID', 'Title', 'Type', 'Priority', 'Category', 'Test Steps', 'Expected Result']
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[1].height = 28

        for i, tc in enumerate(test_cases, start=2):
            steps_text = '\n'.join(tc.get('steps', [])) if isinstance(tc.get('steps'), list) else str(tc.get('steps', ''))
            row_data = [
                tc.get('id', ''),
                tc.get('title', ''),
                tc.get('type', ''),
                tc.get('priority', ''),
                tc.get('category', ''),
                steps_text,
                tc.get('expected_result', ''),
            ]
            ws.append(row_data)
            fill = alt_fill if i % 2 == 0 else PatternFill(fill_type=None)
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=i, column=col_idx)
                if i % 2 == 0:
                    cell.fill = alt_fill
                cell.alignment = Alignment(vertical='top', wrap_text=True)
            ws.row_dimensions[i].height = 60

        col_widths = [10, 35, 14, 12, 20, 60, 45]
        for col_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.freeze_panes = 'A2'

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"test_cases_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename,
        )

    except Exception as exc:
        app.logger.error("download_test_cases_excel error: %s", traceback.format_exc())
        return jsonify({'success': False, 'error': f'Download error: {str(exc)}'}), 500


@app.route('/api/map-excel', methods=['POST'])
def map_excel():
    """
    Map generated test cases against an uploaded Excel test case file.

    multipart/form-data:
        excel_file            – .xlsx / .xls upload
        structured_test_cases – JSON string (array)
    """
    try:
        if 'excel_file' not in request.files:
            return jsonify({'success': False, 'error': 'No Excel file uploaded.'}), 400

        file = request.files['excel_file']
        if not file.filename or not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'error': 'File must be an Excel file (.xlsx or .xls).'}), 400

        raw_cases = request.form.get('structured_test_cases', '[]')
        try:
            generated_cases = json.loads(raw_cases)
            if not isinstance(generated_cases, list):
                raise ValueError("structured_test_cases must be a JSON array")
        except (json.JSONDecodeError, ValueError) as exc:
            return jsonify({'success': False, 'error': f'Invalid structured_test_cases: {exc}'}), 400

        try:
            file_bytes = file.read()
            excel_rows = ExcelProcessor.parse(file_bytes)
        except ExcelParseError as exc:
            return jsonify({'success': False, 'error': str(exc)}), 422

        mapper = ExcelMapper(_bedrock_client)
        result = mapper.map(excel_rows, generated_cases)

        generated_by_id = {tc.get('id', ''): tc for tc in generated_cases}
        new_generated = [generated_by_id[tc_id] for tc_id in result.new_generated_ids if tc_id in generated_by_id]

        return jsonify({
            'success': True,
            'data': {
                'mappings': result.mappings,
                'new_generated': new_generated,
                'stats': result.stats,
            }
        })

    except Exception as exc:
        app.logger.error("map_excel error: %s", traceback.format_exc())
        return jsonify({'success': False, 'error': f'Mapping error: {str(exc)}'}), 500


@app.route('/api/download-mapped-excel', methods=['POST'])
def download_mapped_excel():
    """
    Produce a colour-coded Excel workbook with mapping results.

    multipart/form-data:
        excel_file     – original upload
        mapping_result – JSON string: { mappings, new_generated, stats }
    """
    try:
        if 'excel_file' not in request.files:
            return jsonify({'success': False, 'error': 'No Excel file provided.'}), 400

        file = request.files['excel_file']
        file_bytes = file.read()

        try:
            mapping_data = json.loads(request.form.get('mapping_result', '{}'))
        except json.JSONDecodeError as exc:
            return jsonify({'success': False, 'error': f'Invalid mapping_result JSON: {exc}'}), 400

        try:
            excel_rows = ExcelProcessor.parse(file_bytes)
        except ExcelParseError as exc:
            return jsonify({'success': False, 'error': str(exc)}), 422

        output_bytes = ExcelProcessor.build_output(
            file_bytes,
            excel_rows,
            mapping_data.get('mappings', []),
            mapping_data.get('new_generated', []),
        )

        filename = f"mapped_test_cases_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            io.BytesIO(output_bytes),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename,
        )

    except Exception as exc:
        app.logger.error("download_mapped_excel error: %s", traceback.format_exc())
        return jsonify({'success': False, 'error': f'Download error: {str(exc)}'}), 500


# ─────────────────────────────────────────────
# Health / Jira
# ─────────────────────────────────────────────

@app.route('/health')
def health():
    return jsonify({'status': 'healthy', 'timestamp': datetime.now().isoformat()})


@app.route('/api/jira/health', methods=['GET'])
def jira_health():
    client = ZephyrScaleClient.from_env()
    if client is None:
        return jsonify({'enabled': False, 'connected': False})
    return jsonify({'enabled': True, 'connected': client.health_check()})


if __name__ == '__main__':
    port = int(os.getenv('PORT', 5000))
    debug = os.getenv('FLASK_DEBUG', 'True').lower() == 'true'

    print("=" * 70)
    print("Test Scenario & Coverage Generator")
    print("=" * 70)
    print(f"\nServer starting on http://localhost:{port}")
    print("\nPress Ctrl+C to stop the server")
    print("=" * 70)

    app.run(host='0.0.0.0', port=port, debug=debug)
