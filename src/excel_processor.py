"""
Excel Processor Module
Handles reading existing test case Excel files and producing color-coded mapped output workbooks.
"""

import difflib
import io
from dataclasses import dataclass, field
from datetime import datetime
from typing import Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Canonical internal field names → accepted header aliases (lowercase, stripped)
COLUMN_ALIASES: Dict[str, List[str]] = {
    "test_case": ["test case", "tc id", "test id", "id", "test case id", "tc", "case id"],
    "description": ["description", "test description", "desc", "summary", "test summary"],
    "test_scenario": ["test scenario", "scenario", "test name", "title", "test title"],
    "precondition": ["precondition", "preconditions", "pre-condition", "pre condition", "prerequisites"],
    "test_steps": ["test steps", "steps", "step", "procedure", "test procedure", "test step"],
    "expected_result": ["expected result", "expected", "expected outcome", "result", "expected results"],
}

# openpyxl fill colours for each mapping status
STATUS_FILLS = {
    "MAPPED":         PatternFill(start_color="C8F7C5", end_color="C8F7C5", fill_type="solid"),  # soft green
    "POSSIBLE MATCH": PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),  # amber
    "NOT IMPACTED":   PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid"),  # neutral grey
    "NEW":            PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid"),  # soft blue
}

HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
NEW_SECTION_FILL = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")

MAPPING_COLUMNS = [
    "Mapping Status",
    "Generated TC ID",
    "Generated Title",
    "Match Confidence (%)",
    "QA Notes",
]

# Three decision columns appended after the 5 mapping columns
DECISION_COLUMNS = [
    "Execution Decision",
    "Execution Reason",
    "Suggested Action",
]

# Cell-level fill colours for the Execution Decision column only.
# Row background colours (STATUS_FILLS) are unchanged.
DECISION_FILLS = {
    "MUST_ADD_AND_RUN": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),  # red
    "RUN":              PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),  # green
    "REVIEW":           PatternFill(start_color="FFB300", end_color="FFB300", fill_type="solid"),  # amber
    "SKIP":             PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid"),  # grey (reserved)
}


# ---------------------------------------------------------------------------
# Exceptions
# ---------------------------------------------------------------------------

class ExcelParseError(ValueError):
    """Raised when the uploaded Excel file cannot be parsed meaningfully."""


# ---------------------------------------------------------------------------
# Main class
# ---------------------------------------------------------------------------

class ExcelProcessor:
    """Parse and produce Excel workbooks for test case mapping."""

    # ------------------------------------------------------------------
    # Public API: parse
    # ------------------------------------------------------------------

    @staticmethod
    def parse(file_bytes: bytes) -> List[Dict]:
        """
        Parse an Excel workbook and return a list of test case row dicts.

        Each dict has keys matching COLUMN_ALIASES canonical names plus
        ``row_index`` (1-based worksheet row number) and ``_raw_id``
        (value from the first detected identifier column, for display).

        Raises:
            ExcelParseError: if no recognisable test case columns are found.
        """
        wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ExcelParseError("The uploaded Excel file appears to be empty.")

        # Locate header row: first row containing at least one non-empty cell
        header_row_idx = None
        for i, row in enumerate(rows):
            if any(cell is not None and str(cell).strip() for cell in row):
                header_row_idx = i
                break

        if header_row_idx is None:
            raise ExcelParseError("Could not find a header row in the Excel file.")

        headers = [str(cell).strip().lower() if cell is not None else "" for cell in rows[header_row_idx]]

        # Map each column index → canonical field name
        col_map: Dict[int, str] = {}
        for col_idx, header in enumerate(headers):
            if not header:
                continue
            canonical = ExcelProcessor._match_column(header)
            if canonical and col_idx not in col_map.values():
                col_map[col_idx] = canonical

        if not col_map:
            raise ExcelParseError(
                "No recognisable test case columns found. "
                "Expected headers like: Test Case, Description, Test Scenario, "
                "Precondition, Test Steps, Expected Result."
            )

        # Parse data rows
        result = []
        for row_offset, row in enumerate(rows[header_row_idx + 1:], start=1):
            # Skip entirely empty rows
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            entry: Dict = {"row_index": header_row_idx + 1 + row_offset}
            for col_idx, canonical in col_map.items():
                raw = row[col_idx] if col_idx < len(row) else None
                entry[canonical] = str(raw).strip() if raw is not None else ""

            # Provide defaults for any missing canonical fields
            for canon in COLUMN_ALIASES:
                entry.setdefault(canon, "")

            # Convenience display ID: prefer test_case field, else row number
            entry["_raw_id"] = entry.get("test_case") or f"Row {entry['row_index']}"
            result.append(entry)

        if not result:
            raise ExcelParseError("The Excel file has a header row but no data rows.")

        return result

    # ------------------------------------------------------------------
    # Public API: build_output
    # ------------------------------------------------------------------

    @staticmethod
    def build_output(
        original_bytes: bytes,
        excel_rows: List[Dict],
        mappings: List[Dict],
        new_generated: List[Dict],
    ) -> bytes:
        """
        Produce a colour-coded Excel workbook.

        Appends mapping columns to the right of the original data, colours
        each row by status, and adds a "NEW" section below for generated
        test cases that had no match in the original file.

        Args:
            original_bytes: Raw bytes of the uploaded workbook.
            excel_rows: Parsed rows from ExcelProcessor.parse().
            mappings: List of mapping dicts from MappingResult.mappings.
                      Each item: {excel_row_index, generated_tc_id, status, confidence, notes}
                      generated_tc_id / notes may be None.
            new_generated: List of generated test case dicts (full TC objects) that are NEW.

        Returns:
            bytes of the resulting .xlsx workbook.
        """
        wb = load_workbook(filename=io.BytesIO(original_bytes))
        ws = wb.active

        # Find the last used column in the header row
        header_row_num = ExcelProcessor._detect_header_row_num(ws)
        last_col = ws.max_column or 1

        # Write mapping column headers
        mapping_col_start = last_col + 1
        for offset, col_name in enumerate(MAPPING_COLUMNS):
            cell = ws.cell(row=header_row_num, column=mapping_col_start + offset, value=col_name)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        # Build index: excel row_index → mapping entry
        mapping_by_row: Dict[int, Dict] = {m["excel_row_index"]: m for m in mappings}

        # Apply mapping data to each data row
        for excel_row in excel_rows:
            row_index = excel_row["row_index"]
            mapping = mapping_by_row.get(row_index, {})
            status = mapping.get("status", "NOT IMPACTED")
            fill = STATUS_FILLS.get(status, STATUS_FILLS["NOT IMPACTED"])

            # Colour entire row
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row_index, column=col).fill = fill

            # Write mapping columns
            ws.cell(row=row_index, column=mapping_col_start).value = status
            ws.cell(row=row_index, column=mapping_col_start + 1).value = mapping.get("generated_tc_id") or ""
            ws.cell(row=row_index, column=mapping_col_start + 2).value = mapping.get("generated_title") or ""
            ws.cell(row=row_index, column=mapping_col_start + 3).value = mapping.get("confidence") or ""
            ws.cell(row=row_index, column=mapping_col_start + 4).value = mapping.get("notes") or ""

        # Append NEW generated scenarios section
        if new_generated:
            next_row = ws.max_row + 2  # blank row gap

            # Section header
            separator_cell = ws.cell(row=next_row, column=1, value="NEW — Generated scenarios not in existing test suite")
            separator_cell.fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
            separator_cell.font = Font(color="FFFFFF", bold=True)
            ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=max(last_col + len(MAPPING_COLUMNS), 6))
            next_row += 1

            # NEW rows header
            new_headers = ["TC ID", "Title", "Type", "Priority", "Category", "Steps", "Expected Result"]
            for col_offset, hdr in enumerate(new_headers, start=1):
                cell = ws.cell(row=next_row, column=col_offset, value=hdr)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
            next_row += 1

            for tc in new_generated:
                steps_text = " | ".join(tc.get("steps", [])) if isinstance(tc.get("steps"), list) else str(tc.get("steps", ""))
                row_fill = STATUS_FILLS["NEW"]
                values = [
                    tc.get("id", ""),
                    tc.get("title", ""),
                    tc.get("type", ""),
                    tc.get("priority", ""),
                    tc.get("category", ""),
                    steps_text,
                    tc.get("expected_result", ""),
                ]
                for col_offset, val in enumerate(values, start=1):
                    cell = ws.cell(row=next_row, column=col_offset, value=val)
                    cell.fill = row_fill
                    cell.alignment = Alignment(wrap_text=True)
                next_row += 1

        # Auto-size columns (approximate)
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, min(len(str(cell.value)), 60))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max(12, max_len + 2)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    # ------------------------------------------------------------------
    # Public API: build_decision_output
    # ------------------------------------------------------------------

    @staticmethod
    def build_decision_output(
        original_bytes: bytes,
        excel_rows: List[Dict],
        mappings: List[Dict],
        new_generated: List[Dict],
    ) -> bytes:
        """
        Produce a colour-coded Excel workbook with 8 appended columns:
        the existing 5 mapping columns plus 3 decision columns.

        The 3 decision columns (Execution Decision, Execution Reason,
        Suggested Action) must already be present on each mapping dict —
        call decision_rules.apply_decision() on every item before passing
        the list here.

        Row background colours follow the same STATUS_FILLS convention as
        build_output(). Additionally, the "Execution Decision" cell gets
        its own DECISION_FILLS colour so it stands out visually.

        NEW section rows (unmatched generated TCs) always get
        MUST_ADD_AND_RUN because they have no existing test coverage.

        Args:
            original_bytes: Raw bytes of the uploaded workbook.
            excel_rows:     Parsed rows from ExcelProcessor.parse().
            mappings:       Enriched mapping dicts (status + decision fields).
            new_generated:  Generated TC dicts that are NEW (no Excel match).

        Returns:
            bytes of the resulting .xlsx workbook.
        """
        wb = load_workbook(filename=io.BytesIO(original_bytes))
        ws = wb.active

        header_row_num = ExcelProcessor._detect_header_row_num(ws)
        last_col = ws.max_column or 1

        # Write all 8 appended column headers (5 mapping + 3 decision)
        all_appended = MAPPING_COLUMNS + DECISION_COLUMNS
        mapping_col_start = last_col + 1
        for offset, col_name in enumerate(all_appended):
            cell = ws.cell(row=header_row_num, column=mapping_col_start + offset, value=col_name)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        # Build lookup: excel row_index → enriched mapping dict
        mapping_by_row: Dict[int, Dict] = {m["excel_row_index"]: m for m in mappings}

        for excel_row in excel_rows:
            row_index = excel_row["row_index"]
            mapping = mapping_by_row.get(row_index, {})
            status = mapping.get("status", "NOT IMPACTED")
            fill = STATUS_FILLS.get(status, STATUS_FILLS["NOT IMPACTED"])

            # Colour the entire row with the existing status colour
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row_index, column=col).fill = fill

            # Write the 5 existing mapping columns
            ws.cell(row=row_index, column=mapping_col_start).value     = status
            ws.cell(row=row_index, column=mapping_col_start + 1).value = mapping.get("generated_tc_id") or ""
            ws.cell(row=row_index, column=mapping_col_start + 2).value = mapping.get("generated_title") or ""
            ws.cell(row=row_index, column=mapping_col_start + 3).value = mapping.get("confidence") or ""
            ws.cell(row=row_index, column=mapping_col_start + 4).value = mapping.get("notes") or ""

            # Write the 3 decision columns
            decision = mapping.get("execution_decision", "")
            reason   = mapping.get("execution_reason", "")
            action   = mapping.get("suggested_action", "")

            decision_cell = ws.cell(row=row_index, column=mapping_col_start + 5, value=decision)
            ws.cell(row=row_index, column=mapping_col_start + 6, value=reason)
            ws.cell(row=row_index, column=mapping_col_start + 7, value=action)

            # Colour the Execution Decision cell independently (overrides row fill for that cell)
            if decision in DECISION_FILLS:
                decision_cell.fill = DECISION_FILLS[decision]

        # Append NEW generated scenarios section
        if new_generated:
            next_row = ws.max_row + 2  # blank row gap

            # Section header spanning all columns
            separator_cell = ws.cell(row=next_row, column=1, value="NEW — Generated scenarios not in existing test suite")
            separator_cell.fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
            separator_cell.font = Font(color="FFFFFF", bold=True)
            ws.merge_cells(
                start_row=next_row, start_column=1,
                end_row=next_row, end_column=max(last_col + len(all_appended), 6),
            )
            next_row += 1

            # NEW rows column headers — include the 3 decision columns so they align
            new_headers = ["TC ID", "Title", "Type", "Priority", "Category", "Steps", "Expected Result"]
            for col_offset, hdr in enumerate(new_headers, start=1):
                cell = ws.cell(row=next_row, column=col_offset, value=hdr)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
            # Write decision column headers at the same positions used in the main section
            for offset, hdr in enumerate(DECISION_COLUMNS):
                cell = ws.cell(row=next_row, column=mapping_col_start + 5 + offset, value=hdr)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
            next_row += 1

            for tc in new_generated:
                steps_text = (
                    " | ".join(tc.get("steps", []))
                    if isinstance(tc.get("steps"), list)
                    else str(tc.get("steps", ""))
                )
                row_fill = STATUS_FILLS["NEW"]
                values = [
                    tc.get("id", ""),
                    tc.get("title", ""),
                    tc.get("type", ""),
                    tc.get("priority", ""),
                    tc.get("category", ""),
                    steps_text,
                    tc.get("expected_result", ""),
                ]
                for col_offset, val in enumerate(values, start=1):
                    cell = ws.cell(row=next_row, column=col_offset, value=val)
                    cell.fill = row_fill
                    cell.alignment = Alignment(wrap_text=True)

                # Every NEW TC has no existing coverage — always MUST_ADD_AND_RUN
                decision_cell = ws.cell(
                    row=next_row,
                    column=mapping_col_start + 5,
                    value="MUST_ADD_AND_RUN",
                )
                decision_cell.fill = DECISION_FILLS["MUST_ADD_AND_RUN"]
                ws.cell(
                    row=next_row,
                    column=mapping_col_start + 6,
                    value="No existing test covers this generated scenario. A new test case must be added before release.",
                ).fill = row_fill
                ws.cell(
                    row=next_row,
                    column=mapping_col_start + 7,
                    value="Add New Test",
                ).fill = row_fill

                next_row += 1

        # Auto-size columns (capped at 60 chars to avoid absurdly wide columns)
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, min(len(str(cell.value)), 60))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max(12, max_len + 2)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    # ------------------------------------------------------------------
    # Private helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _match_column(header: str) -> Optional[str]:
        """
        Return the canonical field name for a header string, or None if no match.
        Uses exact match first, then difflib fuzzy match.
        """
        header = header.lower().strip()
        for canonical, aliases in COLUMN_ALIASES.items():
            if header in aliases:
                return canonical

        # Fuzzy fallback across all aliases
        all_aliases = [(alias, canonical) for canonical, aliases in COLUMN_ALIASES.items() for alias in aliases]
        alias_strings = [a for a, _ in all_aliases]
        matches = difflib.get_close_matches(header, alias_strings, n=1, cutoff=0.75)
        if matches:
            for alias, canonical in all_aliases:
                if alias == matches[0]:
                    return canonical
        return None

    @staticmethod
    def _detect_header_row_num(ws) -> int:
        """Return the 1-based row number of the first non-empty row (the header)."""
        for row in ws.iter_rows():
            if any(cell.value for cell in row):
                return row[0].row
        return 1
